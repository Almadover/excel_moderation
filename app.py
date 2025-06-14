from flask import Flask, request, render_template_string
import openpyxl
import io
from datetime import datetime, timedelta

app = Flask(__name__)

# HTML форма для загрузки файла
HTML_FORM = '''
<!doctype html>
<title>Обработка Excel файла</title>
<h2>Загрузите Excel (xlsx) файл</h2>
<form method=post enctype=multipart/form-data>
  <input type=file name=file accept=".xlsx">
  <input type=submit value="Загрузить и обработать">
</form>
'''

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            return "Файл не выбран"
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Найдем индексы нужных столбцов
        header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        try:
            col_resp_idx = header.index('Ответ') + 1  # индекс столбца "Ответ"
            col_time_idx = header.index('Время отправки ответа') + 1  # индекс "Время отправки ответа"
        except ValueError:
            return "Нет нужных столбцов"

        # Обработка значений "Ответ"
        for row in ws.iter_rows(min_row=2, min_col=col_resp_idx, max_col=col_resp_idx):
            cell = row[0]
            if cell.value == 'Можно публиковать':
                cell.value = 'ok'
            elif cell.value == 'Нельзя публиковать':
                cell.value = 'не ok'

        # Обработка времени (добавляем 3 часа, если значение не пустое)
        for row in ws.iter_rows(min_row=2, min_col=col_time_idx, max_col=col_time_idx):
            cell = row[0]
            if cell.value:
                try:
                    # Ваша дата может быть и строчной, и datetime
                    if isinstance(cell.value, str):
                        # Например: '2025-06-14 07:17:04.073000'
                        base_dt = datetime.strptime(cell.value[:19], '%Y-%m-%d %H:%M:%S')
                    else:
                        base_dt = cell.value
                    moscow_dt = base_dt + timedelta(hours=3)
                    cell.value = moscow_dt.strftime('%Y-%m-%d %H:%M:%S')
                except Exception as e:
                    # Оставляем как есть, если не получилось распарсить
                    pass

        # Преобразуем данные в HTML
        from openpyxl.utils import get_column_letter
        table_html = '<table border="1">'
        table_html += '<tr>' + ''.join(f'<th>{cell.value}</th>' for cell in ws[1]) + '</tr>'
        for row in ws.iter_rows(min_row=2):
            row_html = ''
            for cell in row:
                # Если это столбец с 'id Товара', делаем ссылку
                if header[cell.column - 1] == 'id Товара' and cell.value:
                    link = f'https://www.aliexpress.com/item/{cell.value}.html'
                    cell_value = f'<a href="{link}" target="_blank">{cell.value}</a>'
                else:
                    cell_value = '' if cell.value is None else cell.value
                row_html += f'<td>{cell_value}</td>'
            table_html += f'<tr>{row_html}</tr>'
        table_html += '</table>'

        return render_template_string(f'''
            <!doctype html>
            <html>
            <head>
                <title>Результат обработки</title>
            </head>
            <body>
                <h2>Результат обработки файла</h2>
                {table_html}
                <br><a href="/">Назад</a>
            </body>
            </html>
        ''')

    return render_template_string(HTML_FORM)

if __name__ == '__main__':
    app.run(debug=True)